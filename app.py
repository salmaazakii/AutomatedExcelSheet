import openpyxl as excel
from openpyxl.chart import BarChart, Reference

workbook = excel.load_workbook('transactions.xlsx')
sheet = workbook['Sheet1']
# cell = sheet['a1']  # OR cell = sheet.cell(1, 1)
# print(cell.value)
# print(sheet.max_row)

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_value = cell.value * 0.9
    corrected_value_cell = sheet.cell(row, 4)
    corrected_value_cell.value = corrected_value

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

workbook.save('transactions2.xlsx')
